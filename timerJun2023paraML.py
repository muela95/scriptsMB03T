import keyboard
import time
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from colorama import init, Fore

# Initialize colorama
init()
key_press_count = {}


current_directory = os.getcwd()
print("Current working directory:", current_directory)

# Global timer variables
global_timer_running = False
global_timer_start = 0
global_timer_end = 0

# Individual timer variables
timers = {'w': {'running': False, 'start': 0, 'end': 0}, 'a': {'running': False, 'start': 0, 'end': 0},
          's': {'running': False, 'start': 0, 'end': 0}, 'd': {'running': False, 'start': 0, 'end': 0}}

# Create Excel workbook and sheet
wb = Workbook()
ws = wb.active
ws.append(['Beginning', 'End', 'Timer', 'Duration'])

def toggle_timer(timer):
    global global_timer_running, key_press_count

    if global_timer_running:
        if timer not in timers:
            print(f'Invalid timer "{timer}"')
            return

        if timers[timer]['running']:
            # Stop the timer if it is already running
            timers[timer]['running'] = False
            timers[timer]['end'] = time.time() - global_timer_start
            print(f'Timer {timer.upper()} stopped at {timers[timer]["end"]:.2f} seconds.', end='\r')
            
            # Increment key press count for the stopped timer
            if timer not in key_press_count:
                key_press_count[timer] = 1
            else:
                key_press_count[timer] += 1
        else:
            # Start the timer if it is not running
            timers[timer]['running'] = True
            timers[timer]['start'] = time.time() - global_timer_start
            print(f'{Fore.RED}Timer {timer.upper()} started at {timers[timer]["start"]:.2f} seconds.{Fore.RESET}', end='\r')


def on_key_press(event):
    global global_timer_running, global_timer_start, global_timer_end, timers, key_press_count

    if event.name == 'i':
        if not global_timer_running:
            global_timer_start = time.time()
            global_timer_running = True
            print('Global timer started.')

        else:
            print('Global timer is already running.')

    elif event.name in timers.keys():
        if global_timer_running:
            toggle_timer(event.name)

    if event.name == 'space':
        if global_timer_running and not keyboard.is_pressed('i'):
            global_timer_running = False
            global_timer_end = time.time() - global_timer_start  # Calculate global timer duration
            print('Global timer stopped.')

            # Generate filename with the format "yyyymmdd-contaje-hh:mm.xlsx"
            current_datetime = datetime.now().strftime('%Y_%m_%d-%H_%M')
            filename = f'{current_datetime}-contaje.xlsx'

            # Save the .xlsx file to the current working directory
            filepath = os.path.join(current_directory, filename)

            # Write timer data to Excel file
            rows = []
            for timer, data in timers.items():
                if data['start'] != 0 and data['end'] != 0:
                    start_time = data['start']
                    end_time = data['end']
                    duration = end_time - start_time
                    rows.append([start_time, end_time, timer.upper(), duration])

            # Sort rows based on the 'Beginning' column in ascending order
            rows.sort(key=lambda x: x[0])

            for row in rows:
                ws.append(row)

            wb.save(filepath)
            print('Timer data saved to', filepath)
            global_timer_end = time.time() - global_timer_start


            # Calculate the total duration of each timer
            total_durations = {timer: sum([row[3] for row in ws.iter_rows(values_only=True)
                                           if row[2] == timer.upper()])
                               for timer in timers.keys()}
            total_global_duration = global_timer_end
            print('Total Durations:')
            for timer, duration in total_durations.items():
                print(f'{timer.upper()}: {duration:.2f} seconds')

            # Calculate the frequency of each timer
            frequencies = {timer: count for timer, count in key_press_count.items() if timer != 'global'}

            print('Frequency:')
            for timer, frequency in frequencies.items():
                print(f'{timer.upper()}: {frequency:.2f} times')
                key_press_count[timer] = 0
                    
            #key_press_count = {timer: 0 for timer in timers.keys()}
            ws.delete_rows(2, ws.max_row)  # Delete all rows in the worksheet
            # Reset all variables to zero
            global_timer_start = 0
            global_timer_end = 0
            for timer in timers:
                timers[timer]['running'] = False
                timers[timer]['start'] = 0
                timers[timer]['end'] = 0

            print('All variables reset to zero.')


keyboard.on_press(on_key_press)

print('Press letter i to start the global timer.')
print('Press space to stop the global timer and save timer data.')
print('Press w, a, s, or d to toggle individual timers.')

keyboard.wait()
